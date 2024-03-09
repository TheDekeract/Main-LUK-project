import openpyxl
import psycopg2
import re

filename = "8_26_02-02_03.xlsx"

conn = psycopg2.connect(
    dbname="luk",
    user="postgres",
    password="admin",
    host="localhost",
    port="5432"
)
cur = conn.cursor()


def truncate_tables() :  # огромная очистка прошлых расписаний
    cur.execute("TRUNCATE TABLE groups RESTART IDENTITY CASCADE;")
    cur.execute("TRUNCATE TABLE disciplines RESTART IDENTITY CASCADE;")
    cur.execute("TRUNCATE TABLE teachers RESTART IDENTITY CASCADE;")
    cur.execute("TRUNCATE TABLE audiences RESTART IDENTITY CASCADE;")
    cur.execute("TRUNCATE TABLE schedule RESTART IDENTITY CASCADE;")
    print('Clear')


def main_excel() :
    book = openpyxl.load_workbook(filename=filename)
    sheet = book.active

    # cols = sheet.iter_cols(min_col=gr_col, max_col=176, min_row=gr_row, max_row=gr_row)
    start_row = 24  # строка 24 (начальное значение)
    start_col = 9  # столбец I (начальное значение)
    for n in range(15) :
        start_cell = sheet.cell(row=start_row, column=start_col)  # Движение по группам. Только по строке 24
        # print(f"Начальная {n} и ячейка {start_cell}")
        gr_row = start_row
        gr_col = start_col
        for d in range(3) :
            group_cell = sheet.cell(row=gr_row, column=gr_col)  # Движение по группам. Только по строке 24
            day_row = gr_row + 2
            if d == 0 :
                day_col = (gr_col - 2)
            elif d == 1 :
                day_col = (gr_col - 6)
            elif d == 2 :
                day_col = (gr_col - 10)
            for _ in range(6) :  # Выводим расписание на 7 дней
                day_cell = sheet.cell(row=day_row, column=day_col)  # присваиваем ячейку переменной
                time_row = day_row
                time_col = day_col + 1
                for t in range(7) :
                    if time_row == 67 :
                        break
                    time_cell = sheet.cell(row=time_row, column=time_col)  # присваиваем ячейку переменной
                    dis_row = time_row
                    dis_col = gr_col
                    discipline = sheet.cell(row=dis_row, column=dis_col)
                    if discipline.value is not None and "день самостоятельной подготовки" in discipline.value :
                        print("ГОЙДА!")
                        break
                    if discipline.value is not None and "УЧЕБНАЯ ПРАКТИКА" in discipline.value :
                        print("ГОЙДА!")
                        # print("К след дню")
                        # print(f"Содержимое: {discipline.value}, Координаты t: ({dis_row}, {dis_col})")
                        # !
                        teach_row = time_row + 1
                        teach_col = dis_col + 2
                        teacher = sheet.cell(row=teach_row, column=teach_col)
                        # print(f"Содержимое: {teacher.value}, Координаты t: ({teach_row}, {teach_col})")
                        place_row = teach_row
                        place_col = teach_col + 1
                        place = sheet.cell(row=place_row, column=place_col)
                        # print(f"Содержимое: {place.value}, Координаты t: ({place_row}, {place_col})")
                        break
                    else :
                        typeof_row = time_row
                        typeof_col = dis_col + 1
                        type_of_lesson = sheet.cell(row=typeof_row, column=typeof_col)
                        teach_row = time_row
                        teach_col = typeof_col + 1
                        teacher = sheet.cell(row=teach_row, column=teach_col)
                        place_row = time_row
                        place_col = teach_col + 1
                        place = sheet.cell(row=place_row, column=place_col)
                        group_in = group_cell.value
                        print(f"Содержимое: {group_cell.value}, Координаты_: ({gr_row}, {gr_col})")
                        print(f"Содержимое: {day_cell.value}, Координаты_: ({day_row}, {day_col})")
                        print(f"Содержимое: {time_cell.value}, Координаты t: ({time_row}, {time_col})")
                        print(f"Содержимое: {discipline.value}, Координаты t: ({dis_row}, {dis_col})")
                        print(f"Содержимое: {teacher.value}, Координаты t: ({teach_row}, {teach_col})")
                        print(f"Содержимое: {place.value}, Координаты t: ({place_row}, {place_col})")
                        print(f"Содержимое: {type_of_lesson.value}, Координаты t: ({typeof_row}, {typeof_col})")
                        if group_cell is not None :  # отсев ячеек пустых
                            sql = "INSERT INTO groups (name_group) VALUES (%s) ON CONFLICT DO NOTHING"  # добавление группы текущей ячейки в БД
                            cur.execute(sql, (group_in,))  # def для cell_value
                            cur.execute("SELECT id_group FROM groups WHERE name_group = %s", (group_in,))
                            group_id = cur.fetchone()[0]
                            cur.execute("INSERT INTO schedule (id_group) VALUES (%s)", (group_id,))
                        day_in = day_cell.value

                        if day_in is not None :
                            sql = "INSERT INTO schedule (day_of_week_and_date) VALUES (%s) ON CONFLICT DO NOTHING"
                            cur.execute(sql, (day_in,))  # def для cell_value
                        time_in = time_cell.value

                        if time_cell is not None :
                            sql = "INSERT INTO schedule (time_range) VALUES (%s) ON CONFLICT DO NOTHING"
                            cur.execute(sql, (time_in,))  # def для cell_value
                        dis_in = discipline.value

                        if dis_in is not None :  # отсев ячеек пустых
                            sql = "INSERT INTO disciplines (name_discipline) VALUES (%s) ON CONFLICT DO NOTHING"
                            cur.execute(sql, (dis_in,))  # def для cell_value
                            cur.execute("SELECT id_discipline FROM disciplines WHERE name_discipline = %s", (dis_in,))
                            discipline_id = cur.fetchone()[0]
                            cur.execute("INSERT INTO schedule (id_discipline) VALUES (%s)", (discipline_id,))
                        teach_in = teacher.value

                        if teach_in is not None :  # отсев ячеек пустых
                            sql = "INSERT INTO teachers (full_name) VALUES (%s) ON CONFLICT DO NOTHING"
                            cur.execute(sql, (teach_in,))  # def для cell_value
                            cur.execute("SELECT id_teacher FROM teachers WHERE full_name = %s", (teach_in,))
                            teacher_id = cur.fetchone()[0]
                            cur.execute("INSERT INTO schedule (id_teacher) VALUES (%s)", (teacher_id,))
                        placein = place.value

                        if placein is not None :  # отсев ячеек пустых
                            sql = "INSERT INTO audiences (room_number) VALUES (%s) ON CONFLICT DO NOTHING"
                            cur.execute(sql, (placein,))  # def для cell_value
                            cur.execute("SELECT id_audience FROM audiences WHERE room_number = %s", (placein,))
                            audience_id = cur.fetchone()[0]
                            cur.execute("INSERT INTO schedule (id_audience) VALUES (%s)", (audience_id,))
                        typeof_in = type_of_lesson.value

                        if type_of_lesson is not None :  # отсев ячеек пустых
                            sql = "INSERT INTO schedule (lesson_type) VALUES (%s) ON CONFLICT DO NOTHING"
                            cur.execute(sql, (typeof_in,))  # def для cell_value
                        time_row += 1

                    dis_col += 4
                day_row += 7
            day_col -= 4
            gr_col += 4
        start_col += 3


truncate_tables()
main_excel()
conn.commit()
cur.close()
conn.close()
# cell = col[0]  # Получаем ячейку из столбца
# cell_value = cell.value  # присвоение переменной значение текущей ячейки.
#                day_cell = cell.value  # присваиваем переменной значение ячейки
"""        for _ in range(6) :  # Выводим расписание на 7 дней
            day_cell = sheet.cell(row=day_row, column=day_col)  # присваиваем ячейку переменной
            print(f"День недели: {_}, Содержимое: {day_cell.value}, Координаты_: ({day_row}, {day_col})")
            day_cell = cell.value # присваиваем переменной значение ячейки
            if day_cell is not None :# отсев ячеек пустых
                print('Schedule')
                #sql = "INSERT INTO schedule (day_of_week_and_date) VALUES (%s) ON CONFLICT DO NOTHING"
                #cur.execute(sql, (day_cell,))  # def для cell_value
            time_row = 26
            time_col = 8
            for t in range(7) :
                time_cell = sheet.cell(row=time_row, column=time_col)  # присваиваем ячейку переменной
                print(f"Содержимое: {time_cell.value}, Координаты t: ({time_row}, {time_col})")
                time_cell = cell.value # присваиваем переменной значение ячейки
                if time_cell is not None :
                    print('Schedule')
                    #sql = "INSERT INTO schedule (time_range) VALUES (%s) ON CONFLICT DO NOTHING"
                    #cur.execute(sql, (time_cell,))  # def для cell_value
                time_row += 1
            day_row += 7"""
