import openpyxl
import psycopg2

filename = "11_18_03-23_03.xlsx"

conn = psycopg2.connect(
    dbname="luk",
    user="postgres",
    password="admin",
    host="localhost",
    port="5432"
)
cur = conn.cursor()

def truncate_tables() :  # огромная очистка прошлых расписаний
    cur.execute("TRUNCATE TABLE groups RESTART IDENTITY CASCADE;")
    cur.execute("TRUNCATE TABLE disciplines RESTART IDENTITY CASCADE;")
    cur.execute("TRUNCATE TABLE teachers RESTART IDENTITY CASCADE;")
    cur.execute("TRUNCATE TABLE audiences RESTART IDENTITY CASCADE;")
    cur.execute("TRUNCATE TABLE schedule RESTART IDENTITY CASCADE;")
    #cur.execute("TRUNCATE TABLE people RESTART IDENTITY CASCADE;")
    print('Чистая База Данных')

def main_excel() :
    try :
        book = openpyxl.load_workbook(filename=filename)
        sheet = book.active

        start_row = 24  
        start_col = 9  
        gr_row = start_row
        gr_col = start_col
        for n in range(15) :
            start_cell = sheet.cell(row=start_row, column=start_col)  
            if gr_col == 93 :
                gr_col += 6
            for d in range(3) :
                group_cell = sheet.cell(row=gr_row, column=gr_col)  
                if gr_col >= 219 :
                    raise StopIteration
                name_group = group_cell.value
                if name_group is not None and name_group.strip() :  
                    name_group = name_group.strip()
                    day_row = gr_row + 2
                    if d == 0 :
                        day_col = (gr_col - 2)
                    elif d == 1 :
                        day_col = (gr_col - 6)
                    elif d == 2 :
                        day_col = (gr_col - 10)
                for _ in range(6) :  
                    day_cell = sheet.cell(row=day_row, column=day_col)  
                    time_row = day_row
                    time_col = day_col + 1
                    for t in range(7) :
                        if time_row == 67 :
                            break
                        time_cell = sheet.cell(row=time_row, column=time_col)  
                        dis_row = time_row
                        dis_col = gr_col
                        discipline = sheet.cell(row=dis_row, column=dis_col)
                        if discipline.value is not None and "день самостоятельной подготовки" in discipline.value :
                            group_in = group_cell.value
                            if group_cell is not None : 
                                sql = "INSERT INTO groups (name_group) VALUES (%s) ON CONFLICT DO NOTHING"  
                                cur.execute(sql, (group_in,)) 
                                cur.execute("SELECT id_group FROM groups WHERE name_group = %s", (group_in,))
                                group_id = cur.fetchone()[0]
                            day_in = day_cell.value
                            time_in = time_cell.value
                            dis_in = "День самостоятельной подготовки"
                            if dis_in is not None : 
                                sql = "INSERT INTO disciplines (name_discipline) VALUES (%s) ON CONFLICT DO NOTHING"
                                cur.execute(sql, (dis_in,))  
                                cur.execute("SELECT id_discipline FROM disciplines WHERE name_discipline = %s",
                                            (dis_in,))
                                discipline_id = cur.fetchone()[0]
                            teach_in = "!"
                            if teach_in is not None : 
                                sql = "INSERT INTO teachers (full_name) VALUES (%s) ON CONFLICT DO NOTHING"
                                cur.execute(sql, (teach_in,)) 
                                cur.execute("SELECT id_teacher FROM teachers WHERE full_name = %s", (teach_in,))
                                teacher_id = cur.fetchone()[0]
                            placein = "!"
                            if placein is not None :  
                                sql = "INSERT INTO audiences (room_number) VALUES (%s) ON CONFLICT DO NOTHING"
                                cur.execute(sql, (placein,)) 
                                cur.execute("SELECT id_audience FROM audiences WHERE room_number = %s", (placein,))
                                audience_id = cur.fetchone()[0]
                            typeof_in = "!"
                            if typeof_in is not None :
                                sql = """
                                        INSERT INTO schedule (id_group, day_of_week_and_date, time_range, id_discipline, id_teacher, id_audience, lesson_type)
                                        VALUES (%s, %s, %s, %s, %s, %s, %s) ON CONFLICT DO NOTHING;
                                        """
                                cur.execute(sql, (
                                group_id, day_in, time_in, discipline_id, teacher_id, audience_id, typeof_in))
                            break
                        elif discipline.value is not None and "УЧЕБНАЯ ПРАКТИКА" in discipline.value :
                            teach_row = time_row + 1
                            teach_col = dis_col + 2
                            teacher = sheet.cell(row=teach_row, column=teach_col)
                            place_row = teach_row
                            place_col = teach_col + 1
                            place = sheet.cell(row=place_row, column=place_col)
                            group_in = group_cell.value
                            if group_cell is not None :  
                                sql = "INSERT INTO groups (name_group) VALUES (%s) ON CONFLICT DO NOTHING"  # добавление группы текущей ячейки в БД
                                cur.execute(sql, (group_in,))  
                                cur.execute("SELECT id_group FROM groups WHERE name_group = %s", (group_in,))
                                group_id = cur.fetchone()[0]
                            day_in = day_cell.value
                            time_in = time_cell.value
                            dis_in = discipline.value
                            if dis_in is not None :  
                                sql = "INSERT INTO disciplines (name_discipline) VALUES (%s) ON CONFLICT DO NOTHING"
                                cur.execute(sql, (dis_in,))  
                                cur.execute("SELECT id_discipline FROM disciplines WHERE name_discipline = %s",
                                            (dis_in,))
                                discipline_id = cur.fetchone()[0]
                            teach_in = teacher.value
                            if teach_in is not None :  
                                sql = "INSERT INTO teachers (full_name) VALUES (%s) ON CONFLICT DO NOTHING"
                                cur.execute(sql, (teach_in,))  
                                cur.execute("SELECT id_teacher FROM teachers WHERE full_name = %s", (teach_in,))
                                teacher_id = cur.fetchone()[0]
                            placein = place.value
                            if placein is not None :  
                                sql = "INSERT INTO audiences (room_number) VALUES (%s) ON CONFLICT DO NOTHING"
                                cur.execute(sql, (placein,))  
                                cur.execute("SELECT id_audience FROM audiences WHERE room_number = %s", (placein,))
                                audience_id = cur.fetchone()[0]
                            typeof_in = "."
                            if typeof_in is not None :
                                sql = """INSERT INTO schedule (id_group, day_of_week_and_date, time_range, id_discipline, id_teacher, id_audience, lesson_type)VALUES (%s, %s, %s, %s, %s, %s, %s) ON CONFLICT DO NOTHING;"""
                                cur.execute(sql,(group_id, day_in, time_in, discipline_id, teacher_id, audience_id,typeof_in))
                            break
                        elif discipline.value is None:
                            time_row += 1
                        elif discipline.value is not None :
                            typeof_row = time_row
                            typeof_col = dis_col + 1
                            type_of_lesson = sheet.cell(row=typeof_row, column=typeof_col)
                            teach_row = time_row
                            teach_col = typeof_col + 1
                            teacher = sheet.cell(row=teach_row, column=teach_col)
                            place_row = time_row
                            place_col = teach_col + 1
                            place = sheet.cell(row=place_row, column=place_col)
                            group_in = group_cell.value
                            if group_cell is not None : 
                                sql = "INSERT INTO groups (name_group) VALUES (%s) ON CONFLICT DO NOTHING"  
                                cur.execute(sql, (group_in,)) 
                                cur.execute("SELECT id_group FROM groups WHERE name_group = %s", (group_in,))
                                group_id = cur.fetchone()[0]

                            day_in = day_cell.value
                            time_in = time_cell.value
                            dis_in = discipline.value

                            if dis_in is not None : 
                                sql = "INSERT INTO disciplines (name_discipline) VALUES (%s) ON CONFLICT DO NOTHING"
                                cur.execute(sql, (dis_in,))  
                                cur.execute("SELECT id_discipline FROM disciplines WHERE name_discipline = %s",
                                            (dis_in,))
                                discipline_id = cur.fetchone()[0]

                            teach_in = teacher.value

                            if teach_in is not None :  
                                sql = "INSERT INTO teachers (full_name) VALUES (%s) ON CONFLICT DO NOTHING"
                                cur.execute(sql, (teach_in,))  
                                cur.execute("SELECT id_teacher FROM teachers WHERE full_name = %s", (teach_in,))
                                teacher_id = cur.fetchone()[0]

                            placein = place.value

                            if placein is not None : 
                                sql = "INSERT INTO audiences (room_number) VALUES (%s) ON CONFLICT DO NOTHING"
                                cur.execute(sql, (placein,))  
                                cur.execute("SELECT id_audience FROM audiences WHERE room_number = %s", (placein,))
                                audience_id = cur.fetchone()[0]

                            typeof_in = type_of_lesson.value
                            if typeof_in is not None :
                                sql = """
                                        INSERT INTO schedule (id_group, day_of_week_and_date, time_range, id_discipline, id_teacher, id_audience, lesson_type)
                                        VALUES (%s, %s, %s, %s, %s, %s, %s) ON CONFLICT DO NOTHING;
                                        """
                                cur.execute(sql, (
                                group_id, day_in, time_in, discipline_id, teacher_id, audience_id, typeof_in))

                            time_row += 1
                        dis_col += 4
                    day_row += 7
                day_col -= 4
                gr_col += 4

            gr_col += 2
        start_col += 4
    except StopIteration :
        pass

truncate_tables()
main_excel()
conn.commit()
cur.close()
conn.close()
print("База Данных вновь заполнена")
